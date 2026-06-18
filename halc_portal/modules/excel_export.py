"""Excel export — the human review surface (full snapshot + pass1/pass2 deltas).

Builds formatted workbooks with colour-coded scores, green PURSUE highlighting,
frozen header, and an auto-filter table. Human-owned columns (Run Pass 2 /
override score / reason) are merged back from any existing file so regeneration
never wipes reviewer input. Pass 2 recommendation colour takes precedence over
Pass 1 score colour. Column maps (`FULL_COLUMNS`, `PASS1_COLUMNS`,
`PASS2_COLUMNS`) define both the DB->Excel mapping and the column order.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from modules.logutil import log_step

FULL_COLUMNS = [
    ("Tender ID", "tender_id"),
    ("Ref No", "ref_no"),
    ("Center Name", "center_name"),
    ("Category", "category"),
    ("Tender Description", "tender_description"),
    ("Bid Closing Date", "bid_closing_date"),
    ("Bid Opening Date", "bid_opening_date"),
    ("Pass 1 Score", "pass1_score"),
    ("Pass 1 Confidence", "pass1_confidence"),
    ("Pass 1 Domain", "pass1_domain"),
    ("Pass 1 Rationale", "pass1_rationale"),
    ("Pass 2 Score", "pass2_score"),
    ("Pass 2 Confidence", "pass2_confidence"),
    ("Pass 2 Domain", "pass2_domain"),
    ("Pass 2 Rationale", "pass2_rationale"),
    ("Pass 2 Recommendation", "pass2_recommendation"),
    ("Run Pass 2", "run_pass2"),
    ("Human Override Score", "human_override_score"),
    ("Human Override Reason", "human_override_reason"),
    ("Bid Status", "bid_status"),
    ("Previous Closing Date", "previous_closing_date"),
    ("Extension Count", "extension_count"),
    ("First Seen", "first_seen_date"),
    ("Last Seen", "last_seen_at"),
    ("Document URL", "document_url"),
    ("Detail URL", "detail_url"),
    ("Corrigendum URL", "corrigendum_url"),
]

PASS1_COLUMNS = [
    ("Tender ID", "tender_id"),
    ("Ref No", "ref_no"),
    ("Center Name", "center_name"),
    ("Category", "category"),
    ("Tender Description", "tender_description"),
    ("Bid Closing Date", "bid_closing_date"),
    ("Pass 1 Score", "pass1_score"),
    ("Pass 1 Confidence", "pass1_confidence"),
    ("Pass 1 Domain", "pass1_domain"),
    ("Pass 1 Rationale", "pass1_rationale"),
    ("Run Pass 2", "run_pass2"),
    ("Human Override Score", "human_override_score"),
    ("Human Override Reason", "human_override_reason"),
    ("Bid Status", "bid_status"),
    ("Previous Closing Date", "previous_closing_date"),
    ("Extension Count", "extension_count"),
]

PASS2_COLUMNS = [
    ("Tender ID", "tender_id"),
    ("Ref No", "ref_no"),
    ("Center Name", "center_name"),
    ("Category", "category"),
    ("Tender Description", "tender_description"),
    ("Bid Closing Date", "bid_closing_date"),
    ("Pass 1 Score", "pass1_score"),
    ("Pass 2 Score", "pass2_score"),
    ("Pass 2 Confidence", "pass2_confidence"),
    ("Pass 2 Domain", "pass2_domain"),
    ("Pass 2 Rationale", "pass2_rationale"),
    ("Pass 2 Recommendation", "pass2_recommendation"),
    ("Bid Status", "bid_status"),
    ("Previous Closing Date", "previous_closing_date"),
    ("Extension Count", "extension_count"),
]

HUMAN_COLUMNS = ["Run Pass 2", "Human Override Score", "Human Override Reason"]
ID_COL = "Tender ID"

_HEADER_FONT = Font(bold=True)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)
_WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
_NOWRAP_ALIGN = Alignment(wrap_text=False, vertical="top")

_PURSUE_FILL = PatternFill(start_color="70C44E", end_color="70C44E", fill_type="solid")
_SCORE5_FILL = PatternFill(start_color="4DA6FF", end_color="4DA6FF", fill_type="solid")
_SCORE4_FILL = PatternFill(start_color="FFD233", end_color="FFD233", fill_type="solid")
_SCORE3_FILL = PatternFill(start_color="FF9966", end_color="FF9966", fill_type="solid")

_WRAP_COLS = {
    "Tender Description",
    "Pass 1 Rationale",
    "Pass 2 Rationale",
    "Pass 2 Recommendation",
    "Human Override Reason",
}

_WIDTHS = {
    "Tender ID": 22,
    "Ref No": 28,
    "Center Name": 16,
    "Category": 20,
    "Tender Description": 48,
    "Bid Closing Date": 20,
    "Bid Opening Date": 20,
    "Pass 1 Score": 12,
    "Pass 1 Confidence": 16,
    "Pass 1 Domain": 20,
    "Pass 1 Rationale": 55,
    "Pass 2 Score": 12,
    "Pass 2 Confidence": 16,
    "Pass 2 Domain": 20,
    "Pass 2 Rationale": 55,
    "Pass 2 Recommendation": 25,
    "Run Pass 2": 12,
    "Human Override Score": 20,
    "Human Override Reason": 35,
    "Bid Status": 14,
    "Previous Closing Date": 20,
    "Extension Count": 14,
    "First Seen": 14,
    "Last Seen": 20,
    "Document URL": 45,
    "Detail URL": 45,
    "Corrigendum URL": 45,
}

def export_to_excel(rows: list[dict], output_path: str) -> None:
    df = _build_df(rows, FULL_COLUMNS)
    _merge_human_columns(df, output_path)
    df = _sort_df(df)
    _write_excel(df, output_path, "Bids", "BidsTable", hide_closed=True)


def export_pass1_delta(rows: list[dict], output_path: str) -> None:
    _export_delta(rows, output_path, PASS1_COLUMNS, "Pass1", "Pass1Table")


def export_pass2_delta(rows: list[dict], output_path: str) -> None:
    _export_delta(rows, output_path, PASS2_COLUMNS, "Pass2", "Pass2Table")


def _build_df(rows: list[dict], col_map: list[tuple[str, str]]) -> pd.DataFrame:
    excel_cols = [x[0] for x in col_map]
    if not rows:
        return pd.DataFrame(columns=excel_cols)
    data = []
    for row in rows:
        data.append([row.get(db_col) for _, db_col in col_map])
    df = pd.DataFrame(data, columns=excel_cols)
    if "Run Pass 2" in df.columns:
        df["Run Pass 2"] = df["Run Pass 2"].apply(lambda v: "Y" if str(v) == "1" else ("N" if str(v) == "-1" else ""))
    return df


def _merge_human_columns(df: pd.DataFrame, output_path: str) -> None:
    out = Path(output_path)
    if not out.exists() or ID_COL not in df.columns:
        return
    try:
        old = pd.read_excel(out, dtype=str).fillna("")
    except Exception:
        return
    if ID_COL not in old.columns:
        return
    cols = [c for c in [ID_COL, *HUMAN_COLUMNS] if c in old.columns]
    if len(cols) <= 1:
        return
    human = old[cols].drop_duplicates(ID_COL)
    merged = df[[ID_COL]].merge(human, on=ID_COL, how="left")
    for col in HUMAN_COLUMNS:
        if col in df.columns and col in merged.columns:
            old_values = merged[col].fillna("").astype(str)
            new_values = df[col].fillna("").astype(str)
            df[col] = old_values.where(old_values != "", new_values)


def _sort_df(df: pd.DataFrame) -> pd.DataFrame:
    if "Pass 1 Score" in df.columns:
        df["Pass 1 Score"] = pd.to_numeric(df["Pass 1 Score"], errors="coerce")
    sort_cols = [c for c in ["Pass 1 Score", "Bid Closing Date"] if c in df.columns]
    if sort_cols:
        ascending = [False if x == "Pass 1 Score" else True for x in sort_cols]
        df = df.sort_values(by=sort_cols, ascending=ascending, na_position="last").reset_index(drop=True)
    return df


def _export_delta(
    rows: list[dict],
    output_path: str,
    col_map: list[tuple[str, str]],
    sheet_name: str,
    table_name: str,
) -> None:
    df = _build_df(rows, col_map)
    out = Path(output_path)
    expected_cols = [c[0] for c in col_map]
    if out.exists() and not df.empty:
        try:
            old = pd.read_excel(out, dtype=str).fillna("")
            old_cols = list(old.columns)
            # Append only when schema matches; otherwise replace file with new schema.
            if old_cols == expected_cols:
                df = pd.concat([old, df], ignore_index=True)
                if ID_COL in df.columns:
                    df = df.drop_duplicates(subset=[ID_COL], keep="last")
            else:
                log_step(
                    f"{sheet_name}: existing file schema differs; replacing with current delta schema."
                )
        except Exception:
            pass
    df = _sort_df(df)
    _write_excel(df, output_path, sheet_name, table_name, hide_closed=False)


def _write_excel(df: pd.DataFrame, output_path: str, sheet_name: str, table_name: str, hide_closed: bool) -> None:
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        _format_sheet(ws, list(df.columns), table_name)
        if hide_closed and "Bid Status" in df.columns:
            hidden = 0
            for idx, (_, row) in enumerate(df.iterrows(), start=2):
                if str(row.get("Bid Status", "")) == "CLOSED":
                    ws.row_dimensions[idx].hidden = True
                    hidden += 1
            log_step(f"{sheet_name}: {len(df)} row(s), {hidden} CLOSED hidden → {out.name}")
        else:
            log_step(f"{sheet_name}: {len(df)} row(s) → {out.name}")


def _format_sheet(ws, col_names: list[str], table_name: str) -> None:
    n_rows = ws.max_row
    n_cols = ws.max_column
    if n_cols == 0:
        return
    for idx, col_name in enumerate(col_names, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = _WIDTHS.get(col_name, 14)
    ws.row_dimensions[1].height = 30
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
    for r in range(2, n_rows + 1):
        ws.row_dimensions[r].height = 75
        for c, col_name in enumerate(col_names, start=1):
            ws.cell(row=r, column=c).alignment = _WRAP_ALIGN if col_name in _WRAP_COLS else _NOWRAP_ALIGN

    last_col = get_column_letter(n_cols)
    table_ref = f"A1:{last_col}{max(n_rows, 1)}"
    tbl = Table(displayName=table_name, ref=table_ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tbl)

    data_range = f"A2:{last_col}{max(n_rows, 2)}"
    if "Pass 2 Recommendation" in col_names and "Pass 1 Score" in col_names:
        rec_letter = get_column_letter(col_names.index("Pass 2 Recommendation") + 1)
        score_letter = get_column_letter(col_names.index("Pass 1 Score") + 1)
        ws.conditional_formatting.add(
            data_range,
            FormulaRule(formula=[f'OR(${rec_letter}2="PURSUE",${rec_letter}2="PURSUE WITH RAMP-UP")'], fill=_PURSUE_FILL, stopIfTrue=True),
        )
        ws.conditional_formatting.add(data_range, FormulaRule(formula=[f'${rec_letter}2<>""'], stopIfTrue=True))
        ws.conditional_formatting.add(data_range, FormulaRule(formula=[f'${score_letter}2=5'], fill=_SCORE5_FILL))
        ws.conditional_formatting.add(data_range, FormulaRule(formula=[f'${score_letter}2=4'], fill=_SCORE4_FILL))
        ws.conditional_formatting.add(data_range, FormulaRule(formula=[f'${score_letter}2=3'], fill=_SCORE3_FILL))
    elif "Pass 1 Score" in col_names:
        score_letter = get_column_letter(col_names.index("Pass 1 Score") + 1)
        ws.conditional_formatting.add(data_range, FormulaRule(formula=[f'${score_letter}2=5'], fill=_SCORE5_FILL))
        ws.conditional_formatting.add(data_range, FormulaRule(formula=[f'${score_letter}2=4'], fill=_SCORE4_FILL))
        ws.conditional_formatting.add(data_range, FormulaRule(formula=[f'${score_letter}2=3'], fill=_SCORE3_FILL))
    elif "Pass 2 Recommendation" in col_names:
        rec_letter = get_column_letter(col_names.index("Pass 2 Recommendation") + 1)
        ws.conditional_formatting.add(
            data_range,
            FormulaRule(formula=[f'OR(${rec_letter}2="PURSUE",${rec_letter}2="PURSUE WITH RAMP-UP")'], fill=_PURSUE_FILL),
        )
    ws.freeze_panes = "A2"
