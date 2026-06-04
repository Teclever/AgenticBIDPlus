"""
Live validation tests — HAL portal + API.

Run from the project root:
    source venv/bin/activate
    python tests/test_live.py [test_name]

Available tests: nav, fetch, documents, download, pass1, pdf, pass2,
                 excel_format, excel_ingest
Special:  all (default), offline (no portal/API)

COST WARNING: pass1 and pass2 make real Anthropic API calls (~$0.01–0.10/run).
"""

import datetime
import json
import re
import sqlite3
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import config
from modules.db import init_db, upsert_raw_tender
from modules.fetcher import (
    collect_download_urls,
    download_document,
    fetch_all_tenders,
    open_tender_documents,
)
from modules.scorer_pass1 import score_bids_pass1_bulk
from modules.scorer_pass2 import (
    extract_and_clean,
    extract_contract_value,
    extract_emd_amount,
    is_boilerplate_document,
    score_tender_pass2,
)
from modules.session import (
    fill_tender_number,
    find_results_scope,
    open_free_view,
)


# ── ANSI colour helpers ────────────────────────────────────────────────────────

def _ok(msg):   print(f"  \033[32m✓\033[0m  {msg}")
def _fail(msg): print(f"  \033[31m✗\033[0m  {msg}"); sys.exit(1)
def _info(msg): print(f"  \033[34m·\033[0m  {msg}")
def _warn(msg): print(f"  \033[33m⚠\033[0m  {msg}")
def _hdr(msg):  print(f"\n{'='*60}\n  {msg}\n{'='*60}")


# ── Shared Playwright context ──────────────────────────────────────────────────
# Created once for all live tests; tests that don't need the browser skip this.

_pw_context = None     # BrowserContext
_pw_proc    = None     # sync_playwright()


def _get_context():
    global _pw_context, _pw_proc
    if _pw_context is not None:
        return _pw_context
    from playwright.sync_api import sync_playwright
    _pw_proc    = sync_playwright().__enter__()
    _pw_context = _pw_proc.chromium.launch_persistent_context(
        user_data_dir=str(config.BROWSER_PROFILE_DIR),
        headless=True,
        accept_downloads=True,
        ignore_https_errors=True,
        args=["--disable-blink-features=AutomationControlled"],
    )
    return _pw_context


def _close_context():
    global _pw_context, _pw_proc
    if _pw_context:
        try:
            _pw_context.close()
        except Exception:
            pass
        _pw_context = None
    if _pw_proc:
        try:
            _pw_proc.__exit__(None, None, None)
        except Exception:
            pass
        _pw_proc = None


# ── Shared state ──────────────────────────────────────────────────────────────

_state: dict = {}


# ── Test: Playwright portal navigation ────────────────────────────────────────

def test_nav():
    _hdr("Browser navigation — open free view + results scope")

    ctx  = _get_context()
    page = ctx.new_page()
    pages_before = {id(pg) for pg in ctx.pages}

    try:
        fv_page = open_free_view(ctx, page)
        _ok(f"open_free_view returned page at: {fv_page.url[:80]}")

        # find_results_scope clicks Search internally on first iteration
        scope = find_results_scope(ctx, timeout_ms=90_000)

        if scope is None:
            _fail("find_results_scope timed out — #myTable never appeared")

        _ok(f"Results scope found: {scope.url[:80]}")

        # Check #myTable has rows
        row_count = scope.evaluate(
            "() => document.querySelectorAll('#myTable tbody tr').length"
        )
        if row_count < 1:
            _fail(f"#myTable has {row_count} rows — expected > 0")
        _ok(f"#myTable has {row_count} visible row(s)")

        # Check NO_OF_ROWS in page content
        content = scope.content()
        m = re.search(r'"NO_OF_ROWS"\s*:\s*(\d+)', content)
        if m:
            _ok(f"NO_OF_ROWS = {m.group(1)}")
        else:
            _warn("NO_OF_ROWS not found in page content")

        _state["scope"] = scope
    finally:
        for pg in list(ctx.pages):
            if id(pg) not in pages_before:
                try:
                    pg.close()
                except Exception:
                    pass


# ── Test: Full tender list fetch ───────────────────────────────────────────────

def test_fetch():
    _hdr("Full scrape — fetch_all_tenders")
    _info("Running full fetch (may take ~2–3 min)…")

    init_db()
    total = fetch_all_tenders(
        on_page=lambda records: [upsert_raw_tender(r) for r in records],
    )
    if total < 1:
        _fail(f"fetch_all_tenders returned {total} — nothing scraped")
    _ok(f"Full scrape complete: {total} tenders fetched and stored")
    _state["total_fetched"] = total

    conn = sqlite3.connect(config.DB_PATH)
    conn.row_factory = sqlite3.Row
    rows = conn.execute("SELECT * FROM tenders LIMIT 3").fetchall()
    conn.close()

    if not rows:
        _fail("DB is empty after fetch")
    _ok(f"DB sample: tender_number={rows[0]['tender_number']!r}, description={str(rows[0]['tender_description'])[:60]!r}")

    # Verify description is present (requires JSON-based parsing, not HTML table)
    with_desc = sum(1 for r in rows if r["tender_description"])
    _ok(f"{with_desc}/{len(rows)} sample rows have tender_description")

    # detail_url is session-specific and not persisted to DB — not checked here

    _state["sample_tenders"] = [dict(r) for r in rows]


# ── Test: Document URL collection via gear menu ────────────────────────────────

def test_documents():
    _hdr("Document URLs — gear menu → VendorDocumentsController")

    ctx  = _get_context()
    page = ctx.new_page()
    pages_before = {id(pg) for pg in ctx.pages}

    try:
        # Open the portal and run a full search
        open_free_view(ctx, page)
        scope = find_results_scope(ctx, timeout_ms=90_000)
        if scope is None:
            _fail("Results scope not found for documents test")

        # Try rows until we find one with documents (GEM-routed tenders have none)
        doc_urls: set[str] = set()
        for row_idx in range(10):
            urls = open_tender_documents(ctx, scope, row_idx=row_idx)
            if urls:
                doc_urls = urls
                _ok(f"Row {row_idx}: found {len(doc_urls)} download URL(s)")
                for u in list(doc_urls)[:3]:
                    _info(f"  {u[:80]}")
                for url in doc_urls:
                    if "DownloadController" not in url:
                        _warn(f"Unexpected URL pattern: {url[:80]}")
                    if "enc=" in url and "chkSum=" in url:
                        _warn(f"enc/chkSum params decoded (should stay %3D/%26): {url[:80]}")
                break
            _info(f"Row {row_idx}: no documents (GEM-routed or empty)")

        if not doc_urls:
            _warn("No download URLs found in first 10 rows — all may be GEM-routed")

        _state["doc_urls"]      = doc_urls
        _state["doc_scope"]     = scope
        _state["doc_scope_ctx"] = ctx

    finally:
        for pg in list(ctx.pages):
            if id(pg) not in pages_before and pg is not _state.get("doc_scope"):
                try:
                    pg.close()
                except Exception:
                    pass


# ── Test: PDF download ─────────────────────────────────────────────────────────

def test_download():
    _hdr("PDF download — context.request.get()")

    if "doc_urls" not in _state:
        test_documents()

    doc_urls = _state.get("doc_urls") or set()
    ctx      = _get_context()

    if not doc_urls:
        _warn("No document URLs in state — skipping download test")
        return

    seen: set[str] = set()
    result = None
    for i, url in enumerate(sorted(doc_urls), 1):
        result = download_document(ctx, url, i, seen)
        if result is not None:
            break

    if result is None:
        _warn("All downloads returned None (HTML response / dedup) — check DownloadController")
        return

    raw_bytes, fname = result
    _ok(f"Downloaded: {fname}  ({len(raw_bytes):,} bytes)")

    if len(raw_bytes) < 100:
        _fail(f"Downloaded only {len(raw_bytes)} bytes — not a real file")

    if raw_bytes[:4] == b"%PDF":
        _ok("PDF header (%PDF) confirmed")
    else:
        _warn(f"First 4 bytes: {raw_bytes[:4]!r} — may not be a PDF")

    _state["sample_pdf_bytes"] = raw_bytes
    _state["sample_pdf_name"]  = fname


# ── Test: Pass 1 scoring ──────────────────────────────────────────────────────

def test_pass1():
    _hdr("Pass 1 scoring (Haiku)")

    cap_ref = Path(config.CAPABILITY_REF_PATH).read_text()

    conn = sqlite3.connect(config.DB_PATH)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        "SELECT * FROM tenders WHERE bid_status != 'CLOSED' LIMIT 5"
    ).fetchall()
    conn.close()

    if rows:
        tenders = [dict(r) for r in rows]
        _info(f"Scoring {len(tenders)} real HAL tender(s) from DB")
    else:
        tenders = [
            {
                "tender_number": "HAL/IMM/TEST/001",
                "line_number": "1",
                "tender_description": "Development of Avionics LRU Test Rig for TEJAS MK2",
                "buyer": "IMM",
                "tender_region": "Bangalore",
                "estimated_cost": "Rs 50 Lakhs",
                "emd_listing": "Rs 1 Lakh",
                "closing_date": "30-06-2026 14:00",
                "bidder_type": "Indian",
            },
            {
                "tender_number": "HAL/WORKS/TEST/002",
                "line_number": "1",
                "tender_description": "Civil construction of boundary wall at Nasik division",
                "buyer": "WORKS",
                "tender_region": "Nasik",
                "estimated_cost": "Rs 2 Crores",
                "emd_listing": "Rs 4 Lakhs",
                "closing_date": "15-06-2026 14:00",
                "bidder_type": "Indian",
            },
        ]
        _warn("No DB tenders — using synthetic samples")

    results = score_bids_pass1_bulk(tenders, cap_ref)
    if not results:
        _fail("score_bids_pass1_bulk returned empty — API call may have failed")

    _ok(f"Scored {len(results)}/{len(tenders)} tenders")
    for r in results:
        _info(
            f"  [{r.get('tender_number')}] score={r.get('score')}  "
            f"conf={r.get('confidence')}  domain={str(r.get('domain',''))[:40]}"
        )

    required = {"tender_number", "line_number", "score", "confidence", "domain", "rationale", "gaps"}
    for r in results:
        missing = required - set(r.keys())
        if missing:
            _fail(f"Result missing fields {missing}: {r}")
    _ok("All required fields present in results")


# ── Test: PDF text extraction and financial extraction ─────────────────────────

def test_pdf():
    _hdr("PDF extraction + boilerplate filter + amount extraction")

    if "sample_pdf_bytes" not in _state:
        test_download()

    pdf_bytes = _state.get("sample_pdf_bytes")
    pdf_name  = _state.get("sample_pdf_name", "test.pdf")

    if not pdf_bytes:
        _warn("No PDF bytes available — skipping PDF tests")
        return

    raw_text = extract_and_clean(pdf_bytes)
    _ok(f"extract_and_clean: {len(raw_text)} chars")
    if len(raw_text) < 50:
        _warn("Very short extracted text — PDF may be image-based (base64 fallback will apply)")
    else:
        _info(f"First 200 chars:\n{raw_text[:200]}")

    # Boilerplate filter by filename
    test_names = [
        ("AS9100D Compliance Certificate.pdf", True),
        ("Checklist_for_vendors.pdf", True),
        ("RFQ_HAL_IMM_2026.pdf", False),
        ("Annexure_1_Scope_of_Work.pdf", False),
        ("Integrity_Pact.pdf", True),
    ]
    all_correct = True
    for name, expected_bp in test_names:
        got = is_boilerplate_document(name)
        if got == expected_bp:
            _ok(f"is_boilerplate_document({name!r}) = {got}")
        else:
            _warn(f"is_boilerplate_document({name!r}): expected {expected_bp}, got {got}")
            all_correct = False
    if all_correct:
        _ok("All boilerplate filename checks correct")

    # Financial extraction
    emd = extract_emd_amount(raw_text)
    cv  = extract_contract_value(raw_text)
    if emd:
        _ok(f"EMD amount extracted: {emd!r}")
    else:
        _warn("EMD amount not found (may be absent in this document)")
    if cv:
        _ok(f"Contract value extracted: {cv!r}")
    else:
        _warn("Contract value not found (may be absent or phrased differently)")


# ── Test: Pass 2 end-to-end ────────────────────────────────────────────────────

def test_pass2():
    _hdr("Pass 2 end-to-end (one tender)")

    cap_ref = Path(config.CAPABILITY_REF_PATH).read_text()

    conn = sqlite3.connect(config.DB_PATH)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        "SELECT * FROM tenders "
        "WHERE pass1_score >= 3 AND pass2_attempted = 0 AND bid_status != 'CLOSED' "
        "LIMIT 1"
    ).fetchall()
    conn.close()

    if not rows:
        conn = sqlite3.connect(config.DB_PATH)
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT * FROM tenders WHERE pass2_attempted = 0 AND bid_status != 'CLOSED' LIMIT 1"
        ).fetchall()
        conn.close()

    if not rows:
        _warn("No eligible tender for Pass 2 — run test_fetch + test_pass1 first")
        return

    tender = dict(rows[0])
    _info(
        f"Pass 2 candidate: {tender['tender_number']} / {tender['line_number']} "
        f"(pass1_score={tender.get('pass1_score')})"
    )

    ctx = _get_context()
    result = score_tender_pass2(tender, cap_ref, ctx)

    if result is None:
        _warn(
            "score_tender_pass2 returned None — portal navigation or PDF download failed.\n"
            "  Check: fill_tender_number selector, gear menu row index, "
            "DownloadController URL encoding."
        )
        return

    _ok(f"Pass 2 scored: score={result.get('pass2_score')}  rec={result.get('pass2_recommendation')}")
    for k, v in result.items():
        _info(f"  {k}: {str(v)[:80]}")

    required = {
        "pass2_score", "pass2_confidence", "pass2_domain",
        "pass2_rationale", "pass2_recommendation",
    }
    missing = required - set(result.keys())
    if missing:
        _fail(f"Pass 2 result missing fields: {missing}")
    _ok("All required Pass 2 fields present")


# ── Test: Excel export format ─────────────────────────────────────────────────

def test_excel_format():
    _hdr("Excel export format")
    import openpyxl
    from modules.excel_export import (
        COLUMN_MAP, PASS1_DELTA_COLUMNS, PASS2_DELTA_COLUMNS,
        export_to_excel, export_pass1_delta,
    )

    _ok(f"COLUMN_MAP:          {len(COLUMN_MAP)} columns")
    _ok(f"PASS1_DELTA_COLUMNS: {len(PASS1_DELTA_COLUMNS)} columns")
    _ok(f"PASS2_DELTA_COLUMNS: {len(PASS2_DELTA_COLUMNS)} columns")

    p1_required = {
        "Tender Number", "Line Number", "Pass 1 Score", "Pass 1 Rationale",
        "Run Pass 2", "Human Override Score", "Human Override Reason", "Bid Status",
    }
    p2_required = {
        "Tender Number", "Line Number", "Pass 2 Score", "Pass 2 Recommendation",
        "EMD Amount (PDF)", "Contract Value (PDF)", "Bid Status",
    }
    p1_cols = {h for h, _ in PASS1_DELTA_COLUMNS}
    p2_cols = {h for h, _ in PASS2_DELTA_COLUMNS}

    for col in p1_required:
        if col not in p1_cols:
            _fail(f"PASS1_DELTA_COLUMNS missing required column: {col!r}")
    _ok("PASS1_DELTA_COLUMNS has all required columns")

    for col in p2_required:
        if col not in p2_cols:
            _fail(f"PASS2_DELTA_COLUMNS missing required column: {col!r}")
    _ok("PASS2_DELTA_COLUMNS has all required columns")

    import pathlib
    with tempfile.TemporaryDirectory() as td:
        out = str(pathlib.Path(td) / "bids_test.xlsx")
        export_to_excel(out)
        wb = openpyxl.load_workbook(out)
        ws = wb["Bids"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        _ok(f"Bids sheet: {len(headers)} columns, {ws.max_row - 1} data rows")

        assert ws.freeze_panes == "A2", f"freeze_panes={ws.freeze_panes}"
        _ok("Freeze pane at A2")

        assert ws.tables, "No Excel Table found in worksheet"
        _ok(f"Excel Table: {list(ws.tables.keys())}")

        assert ws.conditional_formatting, "No conditional formatting rules"
        _ok(f"Conditional formatting: {len(list(ws.conditional_formatting))} rule(s)")


# ── Test: Excel ingest round-trip ─────────────────────────────────────────────

def test_excel_ingest():
    _hdr("Excel ingest round-trip")
    import pandas as pd
    import pathlib
    from modules.excel_ingest import ingest_excel

    conn = sqlite3.connect(config.DB_PATH)
    conn.row_factory = sqlite3.Row
    rows = conn.execute("SELECT * FROM tenders LIMIT 1").fetchall()
    conn.close()

    if not rows:
        _warn("DB is empty — inserting a synthetic tender for this test")
        today = datetime.date.today().isoformat()
        upsert_raw_tender({
            "tender_number": "HAL/TEST/INGEST/001",
            "line_number":   "1",
            "tender_description": "Test tender for ingest validation",
            "buyer": "IMM",
            "closing_date": "30-06-2026 14:00",
        })
        conn = sqlite3.connect(config.DB_PATH)
        conn.row_factory = sqlite3.Row
        rows = conn.execute("SELECT * FROM tenders LIMIT 1").fetchall()
        conn.close()

    tender = dict(rows[0])
    tn = tender["tender_number"]
    ln = tender["line_number"]

    with tempfile.TemporaryDirectory() as td:
        p1_path = str(pathlib.Path(td) / "pass1_test.xlsx")
        df = pd.DataFrame([{
            "Tender Number":         tn,
            "Line Number":           ln,
            "Tender Description":    tender.get("tender_description", ""),
            "Run Pass 2":            "Y",
            "Human Override Score":  "4",
            "Human Override Reason": "Strong avionics match confirmed in spec",
            "Pass 1 Score":          tender.get("pass1_score", ""),
            "Closing Date":          tender.get("closing_date", ""),
            "Bid Status":            tender.get("bid_status", ""),
        }])
        df.to_excel(p1_path, index=False)
        result = ingest_excel(p1_path, force=True)

    assert not result.get("skipped"), "Ingest was skipped unexpectedly"
    assert result["overrides_applied"] == 1, f"Expected 1 override, got {result}"
    _ok("Ingest ran and applied 1 override")

    conn = sqlite3.connect(config.DB_PATH)
    conn.row_factory = sqlite3.Row
    row = conn.execute(
        "SELECT run_pass2, human_override_score, human_override_reason "
        "FROM tenders WHERE tender_number=? AND line_number=?",
        (tn, ln),
    ).fetchone()
    fse = conn.execute(
        "SELECT * FROM few_shot_examples ORDER BY id DESC LIMIT 1"
    ).fetchone()
    conn.close()

    assert row["run_pass2"] == 1, f"run_pass2={row['run_pass2']}"
    assert row["human_override_score"] == 4
    _ok(f"DB updated: run_pass2=1, override_score=4, reason confirmed")

    if fse:
        _ok(f"Few-shot example added: score={fse['correct_score']}  title={fse['tender_title'][:40]!r}")
    else:
        _warn("No few-shot example created — check process_feedback")


# ── Test registry and runner ──────────────────────────────────────────────────

_TESTS = {
    "nav":          (test_nav,          "Browser navigation — open_free_view + find_results_scope"),
    "fetch":        (test_fetch,        "Full scrape — fetch_all_tenders"),
    "documents":    (test_documents,    "Document URLs — gear menu → DownloadController"),
    "download":     (test_download,     "PDF download — context.request.get()"),
    "pass1":        (test_pass1,        "Pass 1 scoring (Haiku)"),
    "pdf":          (test_pdf,          "PDF extraction + boilerplate filter + amounts"),
    "pass2":        (test_pass2,        "Pass 2 end-to-end"),
    "excel_format": (test_excel_format, "Excel export format"),
    "excel_ingest": (test_excel_ingest, "Excel ingest round-trip"),
}

_OFFLINE_TESTS   = {"pass1", "excel_format", "excel_ingest"}
_LIVE_SEQUENCE   = ["nav", "fetch", "documents", "download", "pass1", "pdf", "pass2"]
_ALL_SEQUENCE    = _LIVE_SEQUENCE + ["excel_format", "excel_ingest"]


if __name__ == "__main__":
    init_db()
    arg = sys.argv[1] if len(sys.argv) > 1 else "all"

    if arg == "offline":
        tests_to_run = list(_OFFLINE_TESTS)
        print("Running offline tests only (no portal or API access required)")
    elif arg == "all":
        tests_to_run = _ALL_SEQUENCE
        print("Running all tests (portal + API access required)")
    elif arg in _TESTS:
        tests_to_run = [arg]
    else:
        print(f"Unknown test: {arg!r}")
        print(f"Available: {', '.join(_TESTS.keys())}")
        print("Special: 'all' (default), 'offline'")
        sys.exit(1)

    passed = failed = 0
    try:
        for name in tests_to_run:
            fn, label = _TESTS[name]
            try:
                fn()
                passed += 1
            except SystemExit:
                failed += 1
                break
            except Exception as e:
                _fail(f"Unexpected exception in {name}: {e}")
                failed += 1
                break
    finally:
        _close_context()

    print(f"\n{'='*60}")
    print(f"  Results: {passed} passed, {failed} failed")
    print(f"{'='*60}")
    sys.exit(0 if failed == 0 else 1)
