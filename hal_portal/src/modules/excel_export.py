import sqlite3

import pandas as pd
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from config import DB_PATH, EXPORTS_DIR

# ── Column maps ────────────────────────────────────────────────────────────────

# Full bids snapshot — all tenders table columns shown to the analyst.
# Order: identity → listing → detail → financial → scoring → human → lifecycle.
COLUMN_MAP = [
    ("Tender Number",         "tender_number"),
    ("Line Number",           "line_number"),
    ("Buyer",                 "buyer"),
    ("Region",                "tender_region"),
    ("Tender Description",    "tender_description"),
    ("Estimated Cost",        "estimated_cost"),
    ("Form Fee",              "form_fee"),
    ("EMD (Listing)",         "emd_listing"),
    ("Tender Stage",          "tender_stage"),
    ("Tender Cover",          "tender_cover"),
    ("Tender Type",           "tender_type"),
    ("Submission Type",       "submission_type"),
    ("Tender For",            "tender_for"),
    ("Bidder Type",           "bidder_type"),
    ("Closing Date",          "closing_date"),
    ("Opening Date",          "opening_date"),
    ("Cost Open Date",        "cost_open_date"),
    ("Announcement Date",     "announcement_date"),
    ("Issue Date From",       "issue_date_from"),
    ("Issue Date To",         "issue_date_to"),
    ("Tender Mode",           "tender_mode"),
    ("Validity of Bid",       "validity_of_bid"),
    ("Contact Person",        "contact_person"),
    ("Contact Email",         "contact_email"),
    ("Qualification Criteria","qualification_criteria"),
    ("Additional Notes",      "additional_notes"),
    ("EMD Amount (PDF)",      "emd_amount"),
    ("Contract Value (PDF)",  "contract_value"),
    ("Pass 1 Score",          "pass1_score"),
    ("Pass 1 Confidence",     "pass1_confidence"),
    ("Pass 1 Domain",         "pass1_domain"),
    ("Pass 1 Rationale",      "pass1_rationale"),
    ("Pass 1 Gaps",           "pass1_gaps"),
    ("Pass 2 Score",          "pass2_score"),
    ("Pass 2 Confidence",     "pass2_confidence"),
    ("Pass 2 Domain",         "pass2_domain"),
    ("Pass 2 Rationale",      "pass2_rationale"),
    ("Pass 2 Gaps",           "pass2_gaps"),
    ("Pass 2 Recommendation", "pass2_recommendation"),
    ("Run Pass 2",            "run_pass2"),
    ("Human Override Score",  "human_override_score"),
    ("Human Override Reason", "human_override_reason"),
    ("Bid Status",            "bid_status"),
    ("Previous Closing Date", "previous_closing_date"),
    ("Extension Count",       "extension_count"),
    ("First Seen",            "first_seen_date"),
    ("Last Updated",          "last_updated_date"),
]

# Columns the human analyst edits in the Excel file; merged back on re-export.
HUMAN_COLUMNS = ["Run Pass 2", "Human Override Score", "Human Override Reason"]

# Pass 1 review file — newly scored tenders; human edits Run Pass 2 + overrides here.
PASS1_DELTA_COLUMNS = [
    ("Tender Number",         "tender_number"),
    ("Line Number",           "line_number"),
    ("Buyer",                 "buyer"),
    ("Region",                "tender_region"),
    ("Tender Description",    "tender_description"),
    ("Estimated Cost",        "estimated_cost"),
    ("Closing Date",          "closing_date"),
    ("Pass 1 Score",          "pass1_score"),
    ("Pass 1 Confidence",     "pass1_confidence"),
    ("Pass 1 Domain",         "pass1_domain"),
    ("Pass 1 Rationale",      "pass1_rationale"),
    ("Pass 2 Score",          "pass2_score"),
    ("Pass 2 Confidence",     "pass2_confidence"),
    ("Pass 2 Domain",         "pass2_domain"),
    ("Pass 2 Rationale",      "pass2_rationale"),
    ("Pass 2 Recommendation", "pass2_recommendation"),
    ("EMD (Listing)",         "emd_listing"),
    ("Run Pass 2",            "run_pass2"),
    ("Human Override Score",  "human_override_score"),
    ("Human Override Reason", "human_override_reason"),
    ("Bid Status",            "bid_status"),
    ("Extension Count",       "extension_count"),
    ("First Seen",            "first_seen_date"),
]

# Pass 2 results file — tenders scored in this pass 2 run.
PASS2_DELTA_COLUMNS = [
    ("Tender Number",         "tender_number"),
    ("Line Number",           "line_number"),
    ("Buyer",                 "buyer"),
    ("Region",                "tender_region"),
    ("Tender Description",    "tender_description"),
    ("Closing Date",          "closing_date"),
    ("Pass 2 Score",          "pass2_score"),
    ("Pass 2 Confidence",     "pass2_confidence"),
    ("Pass 2 Domain",         "pass2_domain"),
    ("Pass 2 Rationale",      "pass2_rationale"),
    ("Pass 2 Recommendation", "pass2_recommendation"),
    ("EMD Amount (PDF)",      "emd_amount"),
    ("Contract Value (PDF)",  "contract_value"),
    ("Bid Status",            "bid_status"),
    ("Extension Count",       "extension_count"),
]

# ── Formatting constants ───────────────────────────────────────────────────────

_COL_WIDTHS = {
    "Tender Number":          28,
    "Line Number":            12,
    "Buyer":                  14,
    "Region":                 22,
    "Tender Description":     45,
    "Estimated Cost":         16,
    "Form Fee":               14,
    "EMD (Listing)":          16,
    "Tender Stage":           16,
    "Tender Cover":           14,
    "Tender Type":            14,
    "Submission Type":        16,
    "Tender For":             14,
    "Bidder Type":            14,
    "Closing Date":           20,
    "Opening Date":           20,
    "Cost Open Date":         18,
    "Announcement Date":      20,
    "Issue Date From":        18,
    "Issue Date To":          18,
    "Tender Mode":            14,
    "Validity of Bid":        16,
    "Contact Person":         22,
    "Contact Email":          28,
    "Qualification Criteria": 45,
    "Additional Notes":       45,
    "EMD Amount (PDF)":       18,
    "Contract Value (PDF)":   20,
    "Pass 1 Score":           13,
    "Pass 1 Confidence":      17,
    "Pass 1 Domain":          22,
    "Pass 1 Rationale":       55,
    "Pass 1 Gaps":            35,
    "Pass 2 Score":           13,
    "Pass 2 Confidence":      17,
    "Pass 2 Domain":          22,
    "Pass 2 Rationale":       60,
    "Pass 2 Gaps":            35,
    "Pass 2 Recommendation":  25,
    "Run Pass 2":             12,
    "Human Override Score":   20,
    "Human Override Reason":  35,
    "Bid Status":             14,
    "Previous Closing Date":  22,
    "Extension Count":        16,
    "First Seen":             14,
    "Last Updated":           16,
}

_WRAP_COLS = {
    "Tender Description",
    "Qualification Criteria",
    "Additional Notes",
    "Pass 1 Rationale",
    "Pass 1 Gaps",
    "Pass 2 Rationale",
    "Pass 2 Gaps",
    "Pass 2 Recommendation",
    "Human Override Reason",
}

_PURSUE_FILL  = PatternFill(start_color="70C44E", end_color="70C44E", fill_type="solid")
_SCORE5_FILL  = PatternFill(start_color="4DA6FF", end_color="4DA6FF", fill_type="solid")
_SCORE4_FILL  = PatternFill(start_color="FFD233", end_color="FFD233", fill_type="solid")
_SCORE3_FILL  = PatternFill(start_color="FF9966", end_color="FF9966", fill_type="solid")
_HEADER_FONT  = Font(bold=True)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)
_WRAP_ALIGN   = Alignment(wrap_text=True,  vertical="top")
_NOWRAP_ALIGN = Alignment(wrap_text=False, vertical="top")


# ── Sheet formatting ───────────────────────────────────────────────────────────

def _apply_sheet_formatting(ws, col_names: list, table_name: str) -> None:
    """Apply column widths, header style, data row height, Excel Table,
    conditional formatting (score colours + PURSUE green), and freeze row 1."""
    n_rows = ws.max_row
    n_cols = ws.max_column

    # Column widths
    for col_idx, col_name in enumerate(col_names, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = _COL_WIDTHS.get(col_name, 14)

    # Header row
    ws.row_dimensions[1].height = 30
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN

    # Data rows
    for row_idx in range(2, n_rows + 1):
        ws.row_dimensions[row_idx].height = 80
        for col_idx, col_name in enumerate(col_names, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = _WRAP_ALIGN if col_name in _WRAP_COLS else _NOWRAP_ALIGN

    # Excel Table (includes autofilter)
    last_col  = get_column_letter(n_cols)
    table_ref = f"A1:{last_col}{max(n_rows, 1)}"
    tbl = Table(displayName=table_name, ref=table_ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tbl)

    # Conditional formatting
    data_range   = f"A2:{last_col}{max(n_rows, 2)}"
    has_p2_rec   = "Pass 2 Recommendation" in col_names
    has_p1_score = "Pass 1 Score" in col_names

    if has_p2_rec and has_p1_score:
        rec_letter   = get_column_letter(col_names.index("Pass 2 Recommendation") + 1)
        score_letter = get_column_letter(col_names.index("Pass 1 Score") + 1)
        # PURSUE → green (stop so P1 colours don't also fire)
        ws.conditional_formatting.add(data_range, FormulaRule(
            formula=[f'OR(${rec_letter}2="PURSUE",${rec_letter}2="PURSUE WITH RAMP-UP")'],
            fill=_PURSUE_FILL, stopIfTrue=True,
        ))
        # Any Pass 2 rec present (non-PURSUE) — stop P1 score rules
        ws.conditional_formatting.add(data_range, FormulaRule(
            formula=[f'${rec_letter}2<>""'], stopIfTrue=True,
        ))
        # P1 score colours — only reached when Pass 2 rec is absent
        ws.conditional_formatting.add(data_range, FormulaRule(
            formula=[f'${score_letter}2=5'], fill=_SCORE5_FILL,
        ))
        ws.conditional_formatting.add(data_range, FormulaRule(
            formula=[f'${score_letter}2=4'], fill=_SCORE4_FILL,
        ))
        ws.conditional_formatting.add(data_range, FormulaRule(
            formula=[f'${score_letter}2=3'], fill=_SCORE3_FILL,
        ))
    elif has_p2_rec:
        rec_letter = get_column_letter(col_names.index("Pass 2 Recommendation") + 1)
        ws.conditional_formatting.add(data_range, FormulaRule(
            formula=[f'OR(${rec_letter}2="PURSUE",${rec_letter}2="PURSUE WITH RAMP-UP")'],
            fill=_PURSUE_FILL,
        ))
    elif has_p1_score:
        score_letter = get_column_letter(col_names.index("Pass 1 Score") + 1)
        ws.conditional_formatting.add(data_range, FormulaRule(
            formula=[f'${score_letter}2=5'], fill=_SCORE5_FILL,
        ))
        ws.conditional_formatting.add(data_range, FormulaRule(
            formula=[f'${score_letter}2=4'], fill=_SCORE4_FILL,
        ))
        ws.conditional_formatting.add(data_range, FormulaRule(
            formula=[f'${score_letter}2=3'], fill=_SCORE3_FILL,
        ))

    ws.freeze_panes = "A2"


# ── Full bids snapshot export ──────────────────────────────────────────────────

def export_to_excel(output_path: str) -> None:
    """Write all tenders to a full-snapshot Excel file.

    If the file already exists, human-editable columns (Run Pass 2,
    Human Override Score, Human Override Reason) are read back and merged
    into the new export so analyst edits are never silently overwritten.
    CLOSED rows are included but hidden by default.
    Sorted: Pass 1 Score DESC, Closing Date ASC.
    """
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)

    # ── Step 1: read back human edits from existing file ──────────────────────
    human_values: dict[tuple, dict] = {}
    import os
    if os.path.exists(output_path):
        try:
            existing_df = pd.read_excel(output_path, dtype=str).fillna("")
            for _, row in existing_df.iterrows():
                tn = str(row.get("Tender Number", "")).strip()
                ln = str(row.get("Line Number",   "")).strip()
                if not tn or not ln:
                    continue
                human_values[(tn, ln)] = {
                    col: str(row.get(col, "")).strip() for col in HUMAN_COLUMNS
                }
        except Exception as e:
            print(f"  [warn] Could not read existing file for human column merge: {e}")

    # ── Step 2: query all tenders ─────────────────────────────────────────────
    sql_cols = [db_col for _, db_col in COLUMN_MAP]
    col_list = ", ".join(sql_cols)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        rows = [dict(r) for r in conn.execute(
            f"SELECT {col_list} FROM tenders"
        ).fetchall()]
    finally:
        conn.close()

    if not rows:
        print("  [info] No tenders in database — empty export.")

    # ── Step 3: build DataFrame ───────────────────────────────────────────────
    excel_cols = [excel_col for excel_col, _ in COLUMN_MAP]
    data = [[row.get(db_col) for _, db_col in COLUMN_MAP] for row in rows]
    df = pd.DataFrame(data, columns=excel_cols)

    # ── Step 4: convert run_pass2 flag → Y / N / "" ───────────────────────────
    df["Run Pass 2"] = df["Run Pass 2"].apply(
        lambda v: "Y" if str(v) == "1" else ("N" if str(v) == "-1" else "")
    )

    # ── Step 5: merge human edits ─────────────────────────────────────────────
    if human_values:
        for col in HUMAN_COLUMNS:
            if col in df.columns:
                df[col] = df[col].astype(object)
        for i, row in df.iterrows():
            key = (str(row["Tender Number"]).strip(), str(row["Line Number"]).strip())
            if key in human_values:
                for col in HUMAN_COLUMNS:
                    saved = human_values[key].get(col, "")
                    if saved:
                        df.at[i, col] = saved

    # ── Step 6: sort ──────────────────────────────────────────────────────────
    df["Pass 1 Score"] = pd.to_numeric(df["Pass 1 Score"], errors="coerce")
    df = df.sort_values(
        by=["Pass 1 Score", "Closing Date"],
        ascending=[False, True],
        na_position="last",
    ).reset_index(drop=True)

    # ── Step 7: write ─────────────────────────────────────────────────────────
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Bids")
        ws = writer.sheets["Bids"]
        _apply_sheet_formatting(ws, excel_cols, "BidsTable")

        # Hide CLOSED rows
        for excel_row, (_, row) in enumerate(df.iterrows(), start=2):
            if str(row.get("Bid Status", "")) == "CLOSED":
                ws.row_dimensions[excel_row].hidden = True

    closed_count = (df["Bid Status"] == "CLOSED").sum()
    print(
        f"  [export] {len(df)} tenders written to {output_path} "
        f"({closed_count} CLOSED rows hidden)"
    )


# ── Delta export ───────────────────────────────────────────────────────────────

def _export_delta(
    tender_list: list,
    col_map: list,
    output_path: str,
    label: str,
    sheet_name: str,
) -> None:
    """Export a subset of tenders to a delta Excel file.

    tender_list — iterable of dicts/Rows, each with tender_number and line_number.
    Appends to an existing same-day file (dedup on composite PK); creates new otherwise.
    """
    if not tender_list:
        print(f"  [export] No {label} tenders to export.")
        return

    # Build composite key list for the WHERE clause
    keys = [
        f"{t['tender_number']}|{t['line_number']}"
        for t in tender_list
    ]
    sql_cols     = [db_col for _, db_col in col_map]
    col_list     = ", ".join(sql_cols)
    placeholders = ",".join("?" * len(keys))

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        rows = [dict(r) for r in conn.execute(
            f"SELECT {col_list} FROM tenders "
            f"WHERE tender_number || '|' || line_number IN ({placeholders})",
            keys,
        ).fetchall()]
    finally:
        conn.close()

    excel_cols = [excel_col for excel_col, _ in col_map]
    data       = [[row.get(db_col) for _, db_col in col_map] for row in rows]
    new_df     = pd.DataFrame(data, columns=excel_cols)

    if "Run Pass 2" in new_df.columns:
        new_df["Run Pass 2"] = new_df["Run Pass 2"].apply(
            lambda v: "Y" if str(v) == "1" else ("N" if str(v) == "-1" else "")
        )

    # Append to existing same-day file, then dedup on composite PK
    import os
    if os.path.exists(output_path):
        try:
            existing_df = pd.read_excel(output_path, dtype=str)
            new_df      = pd.concat([existing_df, new_df], ignore_index=True)
            new_df      = new_df.drop_duplicates(
                subset=["Tender Number", "Line Number"], keep="last"
            )
        except Exception as e:
            print(f"  [warn] Could not read existing {label} file for append: {e}")

    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        new_df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        _apply_sheet_formatting(ws, excel_cols, f"{sheet_name}Table")

    print(f"  [export] {len(new_df)} {label} tenders written to {output_path}")


def export_pass1_delta(tender_list: list, output_path: str) -> None:
    """Export tenders newly scored in this Pass 1 run."""
    _export_delta(tender_list, PASS1_DELTA_COLUMNS, output_path, "Pass1", "Pass1")


def export_pass2_delta(tender_list: list, output_path: str) -> None:
    """Export tenders scored in this Pass 2 run."""
    _export_delta(tender_list, PASS2_DELTA_COLUMNS, output_path, "Pass2", "Pass2")
