"""Eliminator governance loop (S5, chunk 3) — ledger + AI delta + Excel review.

Per ``ELIMINATOR_DESIGN.md`` §7–§9. The live lists are DB rows in ``eliminator_terms``
(seeded once from the mined JSON by :mod:`bidplus.eliminator`); thereafter they change
ONLY through this loop — never a live edit, never model-computed directly into the table:

    runtime ledger  →  periodic AI delta  →  list_change_proposals (staging)
                    →  Excel review on the deploy box (pending → ready → consumed)
                    →  transactional apply to eliminator_terms at run start

Three moving parts:
- **Ledger** (pure code, no AI): a human PROMOTE on a soft-flagged bid is a confirmed
  false-positive (``false_positives++``) and requeues the bid for Pass 1; ACCEPT /
  CLEAR-TABLE on the rest confirms them (``confirmed_rejections++``). A HIGH-SUPPORT term
  is NEVER auto-quarantined (fixed by strengthening the POSITIVE list); only
  low-support/low-precision terms quarantine.
- **AI delta**: emits ADD/REMOVE/REFINE *proposals* (never whole lists), each through a
  deterministic keep-guard, written to ``list_change_proposals`` the moment produced
  (crash-safe) and exported to a risk-colour-coded Excel under ``list_review/pending/``.
- **Apply**: ingest only from ``list_review/ready/`` at run start, reconcile each row to
  approved/rejected, apply approved rows to ``eliminator_terms`` transactionally, mark
  them consumed, move the Excel to ``list_review/consumed/``.
"""

from __future__ import annotations

import datetime
import json
import sqlite3
import uuid
from pathlib import Path

import bidplus.config as config
from bidplus import eliminator

_ADAPTER_PK = {
    "hal": ("tender_number", "line_number"),
    "isro": ("tender_id",),
    "gem": ("bid_number",),
}

# Column that holds the raw bid text in both the tool DB and the parent table —
# must match the column the miner/eliminator runs grams() over.
_ADAPTER_TEXT = {
    "hal": "tender_description",
    "isro": "tender_description",
    "gem": "items",
}

_BID_TEXT_PREVIEW = 600  # chars to include in the delta prompt per bid


def _now() -> str:
    return datetime.datetime.now().isoformat(timespec="seconds")


def _table(portal: str) -> str:
    return f"{portal}_bids"


def _pk_where(portal: str) -> tuple[str, tuple[str, ...]]:
    pk = _ADAPTER_PK[portal]
    return " AND ".join(f"{c}=?" for c in pk), pk


def _split_bid_id(portal: str, bid_id: str) -> list[str]:
    """bid_id is the '|'-joined PK used by the scorer; split it back to PK values."""
    n = len(_ADAPTER_PK[portal])
    return bid_id.split("|") if n > 1 else [bid_id]


def _terms_of(eliminated_by: str | None) -> list[str]:
    if not eliminated_by:
        return []
    return [t.strip() for t in eliminated_by.split(",") if t.strip()]


# ── runtime ledger (pure code, no AI) ───────────────────────────────────────────────

def _bump_stat(parent: sqlite3.Connection, term: str, *, fp: int = 0, cr: int = 0) -> None:
    parent.execute(
        "INSERT INTO eliminator_keyword_stats(term, confirmed_rejections, false_positives, status) "
        "VALUES(?,0,0,'active') ON CONFLICT(term) DO NOTHING", (term,))
    parent.execute(
        "UPDATE eliminator_keyword_stats SET confirmed_rejections=confirmed_rejections+?, "
        "false_positives=false_positives+? WHERE term=?", (cr, fp, term))


def _apply_disposition_rule(parent: sqlite3.Connection, term: str) -> str:
    """Per §7: a HIGH-SUPPORT term wrong on one exception is insufficient-alone, not bad —
    mark 'under_review' (fix via a POSITIVE add), never auto-quarantine. Only
    low-support/low-precision (>=2 FP or precision < 0.98) terms quarantine."""
    row = parent.execute(
        "SELECT confirmed_rejections, false_positives FROM eliminator_keyword_stats WHERE term=?",
        (term,)).fetchone()
    cr, fp = row[0], row[1]
    total = cr + fp
    p = cr / total if total else 1.0
    if cr >= config.ELIM_HIGH_SUPPORT:
        status = "under_review"
    elif fp >= 2 or p < 0.98:
        status = "quarantined"
    else:
        status = "active"
    parent.execute("UPDATE eliminator_keyword_stats SET status=? WHERE term=?", (status, term))
    return status


def promote(parent: sqlite3.Connection, portal: str, bid_id: str, reason: str,
            user_id: int | None = None) -> dict:
    """A human PROMOTE: a confirmed false-elimination. Records false_positives on every
    matched negative term, marks the bid 'promoted' (with reason → feeds the AI delta),
    and REQUEUES it for Pass 1 by clearing its score in the tool DB (the next run's
    eliminator skips a promoted bid, so it reaches Haiku). Does NOT auto-quarantine."""
    eliminator.ensure_schema(parent)
    table = _table(portal)
    where, pk = _pk_where(portal)
    vals = _split_bid_id(portal, bid_id)
    row = parent.execute(
        f"SELECT pass1_eliminated_by, auto_rejected FROM {table} WHERE {where}", vals).fetchone()
    if row is None:
        raise RuntimeError(f"{portal}: no bid {bid_id!r} in {table}")
    terms = _terms_of(row[0])
    for t in terms:
        _bump_stat(parent, t, fp=1)
        _apply_disposition_rule(parent, t)
    parent.execute(
        f"UPDATE {table} SET human_disposition='promoted', human_reason=?, "
        f"disposed_by=?, disposed_at=? WHERE {where}", (reason, user_id, _now(), *vals))
    parent.commit()
    _requeue_in_tool_db(portal, bid_id)
    return {"portal": portal, "bid_id": bid_id, "false_positives_for": terms,
            "statuses": {t: parent.execute(
                "SELECT status FROM eliminator_keyword_stats WHERE term=?", (t,)).fetchone()[0]
                for t in terms}}


def promote_to_five(parent: sqlite3.Connection, portal: str, bid_id: str) -> dict:
    """One-click operator override from the detail page: force Pass-1 score to 5.

    Distinct from promote() (false-elimination governance, which re-queues for
    re-scoring): this is a direct verdict on a bid the model under-rated. The
    pass1_* fields are tool-mirrored, so the override is written to BOTH the
    parent and the tool DB — parent-only would silently revert on the next
    nightly merge. The '[operator-promoted]' rationale prefix is the durable
    marker the UI renders a badge from. Score-5 unsummarized bids are picked up
    by the next nightly Sonnet pass automatically."""
    table = _table(portal)
    where, _pk = _pk_where(portal)
    vals = _split_bid_id(portal, bid_id)
    row = parent.execute(
        f"SELECT pass1_score, pass1_rationale FROM {table} WHERE {where}", vals).fetchone()
    if row is None:
        raise LookupError(f"{portal}: no bid {bid_id!r} in {table}")
    old_score, rationale = row
    rationale = rationale or ""
    if not rationale.startswith("[operator-promoted]"):
        rationale = f"[operator-promoted] {rationale}".strip()
    sets = ("pass1_score=5, pass1_method='model', pass1_rationale=?, "
            "pass1_eliminated_by=NULL, auto_rejected=0")
    parent.execute(f"UPDATE {table} SET {sets} WHERE {where}", (rationale, *vals))
    parent.commit()

    from bidplus.adapters.gem import GeMAdapter
    from bidplus.adapters.hal import HALAdapter
    from bidplus.adapters.isro import ISROAdapter
    adapter = {"hal": HALAdapter, "isro": ISROAdapter, "gem": GeMAdapter}[portal]()
    spec = adapter._SCORING
    tool_where = " AND ".join(f"{c}=?" for c in spec["pk"])
    conn = sqlite3.connect(adapter.tool_db_path())
    try:
        conn.execute(f"UPDATE {spec['table']} SET {sets} WHERE {tool_where}",
                     (rationale, *vals))
        conn.commit()
    finally:
        conn.close()
    return {"portal": portal, "bid_id": bid_id, "old_score": old_score, "new_score": 5}


def accept(parent: sqlite3.Connection, portal: str, bid_ids: list[str] | None = None) -> dict:
    """ACCEPT / CLEAR-TABLE: every still-undisposed auto-rejected bid is a confirmed-correct
    rejection (``confirmed_rejections++`` on its matched terms). ``bid_ids=None`` clears the
    whole portal table (all undisposed auto-rejected bids)."""
    eliminator.ensure_schema(parent)
    table = _table(portal)
    where, pk = _pk_where(portal)
    if bid_ids is None:
        rows = parent.execute(
            f"SELECT {', '.join(pk)}, pass1_eliminated_by FROM {table} "
            f"WHERE auto_rejected=1 AND human_disposition IS NULL").fetchall()
    else:
        rows = []
        for bid_id in bid_ids:
            r = parent.execute(
                f"SELECT {', '.join(pk)}, pass1_eliminated_by FROM {table} "
                f"WHERE {where} AND auto_rejected=1 AND human_disposition IS NULL",
                _split_bid_id(portal, bid_id)).fetchone()
            if r is not None:
                rows.append(r)
    n_pk = len(pk)
    confirmed = 0
    for r in rows:
        pkvals = list(r[:n_pk])
        for t in _terms_of(r[n_pk]):
            _bump_stat(parent, t, cr=1)
        parent.execute(
            f"UPDATE {table} SET human_disposition='accepted', disposed_at=? WHERE {where}",
            (_now(), *pkvals))
        confirmed += 1
    parent.commit()
    return {"portal": portal, "accepted": confirmed}


def _requeue_in_tool_db(portal: str, bid_id: str) -> None:
    """Clear a promoted bid's Pass-1 fields in its tool DB so the next centralized run
    re-scores it (the eliminator skips promoted bids → it reaches Haiku)."""
    from bidplus.adapters.gem import GeMAdapter
    from bidplus.adapters.hal import HALAdapter
    from bidplus.adapters.isro import ISROAdapter
    adapter = {"hal": HALAdapter, "isro": ISROAdapter, "gem": GeMAdapter}[portal]()
    spec = adapter._SCORING
    where = " AND ".join(f"{c}=?" for c in spec["pk"])
    conn = sqlite3.connect(adapter.tool_db_path())
    try:
        conn.execute(
            f"UPDATE {spec['table']} SET pass1_score=NULL, pass1_method=NULL, "
            f"pass1_eliminated_by=NULL, auto_rejected=0 WHERE {where}",
            _split_bid_id(portal, bid_id))
        conn.commit()
    finally:
        conn.close()


def promoted_bid_ids(parent: sqlite3.Connection, portal: str) -> set[str]:
    """bid_ids a human has promoted — the centralized scorer must send these to Haiku
    (skip the eliminator) so a confirmed in-scope bid is never re-eliminated."""
    table = _table(portal)
    pk = _ADAPTER_PK[portal]
    try:
        rows = parent.execute(
            f"SELECT {', '.join(pk)} FROM {table} WHERE human_disposition='promoted'").fetchall()
    except sqlite3.OperationalError:
        return set()
    return {"|".join("" if v is None else str(v) for v in r) for r in rows}


# ── periodic AI delta (ELIMINATOR_DESIGN §8) ────────────────────────────────────────

def _last_batch_at(parent: sqlite3.Connection) -> str | None:
    r = parent.execute("SELECT MAX(created_at) FROM list_change_proposals").fetchone()
    return r[0] if r else None


def pending_promotions(parent: sqlite3.Connection) -> list[dict]:
    """Promotion reasons accrued since the last AI delta batch (drive the next one)."""
    cutoff = _last_batch_at(parent)
    out: list[dict] = []
    for portal in config.PORTALS:
        table = _table(portal)
        pk = _ADAPTER_PK[portal]
        try:
            text_col = _ADAPTER_TEXT[portal]
            q = (f"SELECT {', '.join(pk)}, pass1_eliminated_by, human_reason, disposed_at, "
                 f"{text_col} "
                 f"FROM {table} WHERE human_disposition='promoted' AND human_reason IS NOT NULL")
            args: tuple = ()
            if cutoff:
                q += " AND (disposed_at IS NULL OR disposed_at > ?)"
                args = (cutoff,)
            for r in parent.execute(q, args).fetchall():
                raw_text = r[len(pk) + 3] or ""
                out.append({"portal": portal,
                            "bid_id": "|".join("" if v is None else str(v) for v in r[:len(pk)]),
                            "eliminated_by": r[len(pk)], "reason": r[len(pk) + 1],
                            "bid_text": raw_text[:_BID_TEXT_PREVIEW]})
        except sqlite3.OperationalError:
            continue
    return out


def should_run_delta(parent: sqlite3.Connection) -> bool:
    """Fire when accumulated promotion-reasons reach the threshold OR a week has passed
    since the last batch (whichever first) — and only if there is something to learn from."""
    promos = pending_promotions(parent)
    if not promos:
        return False
    if len(promos) >= config.ELIM_DELTA_PROMOTION_THRESHOLD:
        return True
    last = _last_batch_at(parent)
    if last is None:
        return True
    age = datetime.datetime.now() - datetime.datetime.fromisoformat(last)
    return age.days >= config.ELIM_DELTA_WEEKLY_DAYS


def _score_ge3_grams(parent: sqlite3.Connection, terms: eliminator.Terms) -> set[str]:
    """Every gram present in a bid scored >= 3 — the keep-guard set: a proposed NEGATIVE
    add hitting any of these could start killing pursuable bids and is blocked."""
    from bidplus.adapters.gem import GeMAdapter
    from bidplus.adapters.hal import HALAdapter
    from bidplus.adapters.isro import ISROAdapter
    specs = {"hal": HALAdapter, "isro": ISROAdapter, "gem": GeMAdapter}
    grams: set[str] = set()
    for portal, cls in specs.items():
        text_col = cls._SCORING["text"]
        table = _table(portal)
        try:
            rows = parent.execute(
                f"SELECT {text_col} FROM {table} WHERE pass1_score >= 3").fetchall()
        except sqlite3.OperationalError:
            continue
        for (txt,) in rows:
            grams |= eliminator._grams(txt or "", terms.stop)
    return grams


def _current_lists(parent: sqlite3.Connection) -> dict:
    t = eliminator.load_terms(parent)
    return {"neg_phrases": sorted(t.neg_phrases), "neg_words": sorted(t.neg_words),
            "pos_phrases": sorted(t.pos_phrases), "pos_tokens": sorted(t.pos_tokens)}


def _raw_governance(prompt: str) -> str:
    import anthropic
    client = anthropic.Anthropic(api_key=config.require_api_key())
    resp = client.messages.create(
        model=config.GOVERNANCE_MODEL, max_tokens=4096,
        messages=[{"role": "user", "content": prompt}])
    return resp.content[0].text


def _build_delta_prompt(parent: sqlite3.Connection, promos: list[dict]) -> str:
    lists = _current_lists(parent)
    stats = parent.execute(
        "SELECT term, confirmed_rejections, false_positives, status "
        "FROM eliminator_keyword_stats ORDER BY false_positives DESC, confirmed_rejections DESC "
        "LIMIT 60").fetchall()
    ledger = [{"term": s[0], "confirmed": s[1], "false_positives": s[2], "status": s[3]}
              for s in stats]
    return (
        "You maintain a two-pass keyword eliminator for government bid scoring.\n"
        "eliminate = negative_hit AND NOT positive_hit. The NEGATIVE list removes obvious "
        "junk before an LLM scores it; the POSITIVE list rescues in-scope bids that trip a "
        "negative term.\n\n"
        "RULES:\n"
        "- Output a DELTA only — a JSON array of small changes, never whole lists.\n"
        "- Prefer fixing a false-elimination by ADDING a high-precision POSITIVE term, not by "
        "removing a negative term (removal re-admits all that term's junk).\n"
        "- Positive terms must be HIGH-PRECISION (never generic: software/test/system/data/"
        "service/development).\n"
        "- A negative ADD must not match any pursuable bid.\n\n"
        f"CURRENT LISTS:\n{json.dumps(lists)}\n\n"
        f"LEDGER (negative-term stats):\n{json.dumps(ledger)}\n\n"
        f"NEW HUMAN PROMOTIONS (false-eliminations to learn from):\n{json.dumps(promos)}\n\n"
        "Each promotion includes: eliminated_by (the negative term that fired), reason (the "
        "human's explanation of why the bid is in-scope), and bid_text (the first ~600 chars "
        "of the actual bid — use this to identify precise positive rescue terms or to judge "
        "whether the negative term needs refining).\n\n"
        "Return ONLY a JSON array. Each element: {list_type:'neg'|'pos', change_type:'add'|"
        "'remove'|'refine', term:str, kind:'phrase'|'word'|'token', refine_to:str|null, "
        "ai_rationale:str, source_reason:str}."
    )


def _parse_delta(text: str) -> list[dict]:
    import re
    m = re.search(r"\[.*\]", text or "", re.DOTALL)
    if not m:
        return []
    try:
        arr = json.loads(m.group())
    except json.JSONDecodeError:
        return []
    return [d for d in arr if isinstance(d, dict) and d.get("term") and d.get("change_type")]


def _risk_of(d: dict) -> str:
    """§9.3 — colour by RISK not direction. HIGH = could harm correctness."""
    lt, ct = d.get("list_type"), d.get("change_type")
    if (lt == "neg" and ct == "add") or (lt == "pos" and ct == "remove"):
        return "high"
    return "low"


def generate_delta(parent: sqlite3.Connection, call_fn=None, export: bool = True) -> dict:
    """Run the AI delta: build the prompt, get ADD/REMOVE/REFINE proposals, pass each through
    the deterministic keep-guard, persist survivors to ``list_change_proposals`` (status
    'proposed') and export a colour-coded Excel to ``list_review/pending/``. Returns a summary.
    """
    eliminator.ensure_schema(parent)
    promos = pending_promotions(parent)
    call_fn = call_fn or _raw_governance
    raw = call_fn(_build_delta_prompt(parent, promos))
    proposals = _parse_delta(raw)

    terms = eliminator.load_terms(parent)
    guard_grams = _score_ge3_grams(parent, terms)
    batch_id = uuid.uuid4().hex[:12]
    now = _now()
    kept = blocked = 0
    for d in proposals:
        lt = d.get("list_type"); ct = d.get("change_type"); term = (d.get("term") or "").lower().strip()
        if not term or lt not in ("neg", "pos") or ct not in ("add", "remove", "refine"):
            continue
        # keep-guard: a negative ADD that hits a pursuable (score>=3) bid is blocked outright.
        if lt == "neg" and ct == "add" and term in guard_grams:
            blocked += 1
            continue
        # never resolve a promotion by removing an under_review high-support negative term.
        if lt == "neg" and ct == "remove":
            st = parent.execute(
                "SELECT status FROM eliminator_keyword_stats WHERE term=?", (term,)).fetchone()
            if st and st[0] == "under_review":
                blocked += 1
                continue
        parent.execute(
            "INSERT INTO list_change_proposals(batch_id, list_type, term, kind, change_type, "
            "refine_to, risk, ai_rationale, source_reasons, status, created_at) "
            "VALUES(?,?,?,?,?,?,?,?,?,'proposed',?)",
            (batch_id, lt, term, d.get("kind"), ct, d.get("refine_to"), _risk_of(d),
             d.get("ai_rationale"), d.get("source_reason"), now))
        kept += 1
    parent.commit()

    path = None
    if export and kept:
        path = export_proposals_excel(parent, batch_id)
    return {"batch_id": batch_id, "proposed": len(proposals), "kept": kept,
            "blocked_by_guard": blocked, "promotions_used": len(promos),
            "excel": str(path) if path else None}


# ── Excel review surface (ELIMINATOR_DESIGN §9) ─────────────────────────────────────

def list_review_dir(sub: str = "") -> Path:
    d = config.RUNTIME_DIR / "list_review" / sub if sub else config.RUNTIME_DIR / "list_review"
    d.mkdir(parents=True, exist_ok=True)
    return d


_EXCEL_COLS = ["id", "list_type", "change_type", "term", "kind", "refine_to", "risk",
               "ai_rationale", "source_reasons", "REJECT? (blank=approve)"]


def export_proposals_excel(parent: sqlite3.Connection, batch_id: str) -> Path:
    """Write the batch's 'proposed' rows to a colour-coded Excel under ``pending/`` —
    a regenerable VIEW of the staging table (HIGH-risk red, on top). The staging table,
    not this file, is the durable home (so a deleted/edited Excel loses nothing but marks)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    rows = parent.execute(
        "SELECT id, list_type, change_type, term, kind, refine_to, risk, ai_rationale, "
        "source_reasons FROM list_change_proposals WHERE batch_id=? AND status='proposed' "
        "ORDER BY CASE risk WHEN 'high' THEN 0 ELSE 1 END, list_type, term", (batch_id,)).fetchall()

    wb = Workbook(); ws = wb.active; ws.title = "list_changes"
    ws.append(_EXCEL_COLS)
    for c in ws[1]:
        c.font = Font(bold=True)
    red = PatternFill("solid", fgColor="FFC7CE")
    for r in rows:
        ws.append(list(r) + [""])
        if r[6] == "high":
            for c in ws[ws.max_row]:
                c.fill = red
    path = list_review_dir("pending") / f"list_changes_{batch_id}.xlsx"
    wb.save(path)
    return path


def _apply_proposal(parent: sqlite3.Connection, p: sqlite3.Row) -> None:
    """Apply one approved proposal to ``eliminator_terms`` (source='ai')."""
    lt, ct, term, kind, refine_to = p["list_type"], p["change_type"], p["term"], p["kind"], p["refine_to"]
    now = _now()
    if ct == "add":
        parent.execute(
            "INSERT INTO eliminator_terms(list_type, kind, term, is_guarded, active, source, "
            "rationale, created_at) VALUES(?,?,?,0,1,'ai',?,?) "
            "ON CONFLICT(list_type, kind, term) DO UPDATE SET active=1",
            (lt, kind, term, p["ai_rationale"], now))
    elif ct == "remove":
        parent.execute(
            "UPDATE eliminator_terms SET active=0 WHERE list_type=? AND term=?", (lt, term))
    elif ct == "refine" and refine_to:
        parent.execute(
            "UPDATE eliminator_terms SET active=0 WHERE list_type=? AND term=?", (lt, term))
        parent.execute(
            "INSERT INTO eliminator_terms(list_type, kind, term, is_guarded, active, source, "
            "rationale, created_at) VALUES(?,?,?,0,1,'ai',?,?) "
            "ON CONFLICT(list_type, kind, term) DO UPDATE SET active=1",
            (lt, kind, refine_to.lower().strip(), p["ai_rationale"], now))


def ingest_ready(parent: sqlite3.Connection) -> dict:
    """Run-start step: ingest ONLY from ``list_review/ready/`` (a mid-edit file in
    ``pending/`` is never read). Reconcile each row → approved/rejected, apply approved
    rows to ``eliminator_terms`` TRANSACTIONALLY, mark consumed, move the Excel to
    ``consumed/``. Idempotent: a file with no matching 'proposed' rows is a no-op move."""
    from openpyxl import load_workbook

    eliminator.ensure_schema(parent)
    ready = list_review_dir("ready")
    applied = rejected = files = 0
    for xlsx in sorted(ready.glob("*.xlsx")):
        if xlsx.name.startswith("~$"):
            continue
        wb = load_workbook(xlsx, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            continue
        hdr = {str(h).strip().lower(): i for i, h in enumerate(rows[0]) if h is not None}
        id_i = hdr.get("id")
        rej_i = next((i for k, i in hdr.items() if k.startswith("reject")), None)
        if id_i is None:
            continue
        try:
            parent.execute("BEGIN")
            for r in rows[1:]:
                if id_i >= len(r) or r[id_i] in (None, ""):
                    continue
                pid = int(r[id_i])
                p = parent.execute(
                    "SELECT * FROM list_change_proposals WHERE id=? AND status='proposed'",
                    (pid,)).fetchone()
                if p is None:
                    continue
                rejected_mark = rej_i is not None and rej_i < len(r) and str(r[rej_i] or "").strip() != ""
                if rejected_mark:
                    parent.execute(
                        "UPDATE list_change_proposals SET status='rejected', decided_at=? WHERE id=?",
                        (_now(), pid))
                    rejected += 1
                else:
                    _apply_proposal(parent, p)
                    parent.execute(
                        "UPDATE list_change_proposals SET status='approved', decided_at=? WHERE id=?",
                        (_now(), pid))
                    applied += 1
            # everything we touched in this batch_id that is now decided → consumed
            parent.execute(
                "UPDATE list_change_proposals SET status='consumed' "
                "WHERE status IN ('approved','rejected') AND decided_at IS NOT NULL")
            parent.commit()
        except Exception:
            parent.rollback()
            raise
        dest = list_review_dir("consumed") / xlsx.name
        xlsx.replace(dest)
        files += 1
    return {"files": files, "applied": applied, "rejected": rejected}
