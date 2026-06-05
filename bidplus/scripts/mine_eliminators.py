"""Pre-Pass-1 eliminator miner (and scored-corpus exporter) — protect-≥3.

Replaces the old GeM exclusion pre-filter with a single, portal-agnostic,
high-precision keyword gate mined from ALL historically-scored bids across the
three tools. A bid whose text matches a "safe" keyword is assigned score 0
WITHOUT a model call; everything else goes to Pass-1 (Haiku) as usual. This is a
cheap cost-skimmer, NOT a classifier — it only removes obvious junk.

Three artifacts are produced:
  1. $BIDPLUS_RUNTIME_DIR/exports/scored_corpus.xlsx — every scored bid
     (portal, bid_id, score, text) in one sheet, so the keyword list can be
     regenerated with other tools (Excel / pandas / an LLM).
  2. bidplus/data/eliminator_keywords.json — the curated keyword list the
     runtime gate consumes. MACHINE-GENERATED: review before enabling.
  3. $BIDPLUS_RUNTIME_DIR/exports/score2_review.xlsx — the score-2 bids this set
     would eliminate (under the two-signal rule), for the reviewer to eyeball.

Safety model — "protect ≥3":
  reject  = bids scored 0 or 1              (what we want to eliminate)
  protect = bids scored >= PROTECT_FROM (3) (NEVER eliminate — pursuable)
  tolerate = bids scored == 2               (knowingly droppable borderlines)
  A gram qualifies when reject_support >= MIN_SUPPORT AND it has ZERO collisions
  in the protect band (scores 3/4/5). Score-2 collisions are TOLERATED and
  tracked per keyword (keep_s2_per_keyword) as the reviewer's budget lever.

Guarded unigrams (two-signal): any qualifying unigram that whole-word-matches a
term in capability_reference.md is in-scope vocabulary and must NOT eliminate on
its own. At runtime it contributes only under the two-signal rule (the bid also
has a phrase hit or >= 2 distinct unigram hits). A lone `liquid` never cuts.
GUARDED_NOISE (ordinary English the capability doc happens to contain) is trimmed
out of the active set entirely into excluded_generic_unigrams.

Run:  python -m bidplus.scripts.mine_eliminators
"""

from __future__ import annotations

import datetime as _dt
import json
import os
import re
import sqlite3
from collections import Counter
from pathlib import Path

from openpyxl import Workbook

# Repo root: bidplus/scripts/mine_eliminators.py -> parents[2]
_REPO_ROOT = Path(__file__).resolve().parents[2]
_DATA_DIR = _REPO_ROOT / "bidplus" / "data"
_JSON_OUT = _DATA_DIR / "eliminator_keywords.json"
_CAP_REF = _DATA_DIR / "capability_reference.md"

# Per-portal historical (in-tree) corpus: db path, table, text col, PK cols.
SOURCES = {
    "hal":  ("hal_portal/data/bids.db",  "tenders", "tender_description", ["tender_number", "line_number"]),
    "isro": ("isro_portal/data/bids.db", "bids",    "tender_description", ["tender_id"]),
    "gem":  ("gem_portal/data/bids.db",  "bids",    "items",              ["bid_number"]),
}

# ── tuning (locked defaults) ────────────────────────────────────────────────
MIN_SUPPORT = 20    # a keyword must hit >= this many reject (0/1) bids
PROTECT_FROM = 3    # never eliminate on a gram seen in any bid scored >= this
                    # (score 2 is tolerated and tracked, not protected)

# Tokens dropped before forming grams (units / filler that carry no signal).
STOPWORDS = set(
    "for the and with of to in on at by are per nos set qty quantity supply "
    "supplying providing provision various item items make model type size new "
    "used non com www http".split()
)

# Guarded unigrams that are ordinary English the capability doc merely contains —
# trimmed out of the active set entirely (held for review). The remaining
# capability collisions (liquid, structure, tower, rack, lift, life, weight,
# heavy, disposable, …) stay guarded behind the two-signal rule.
GUARDED_NOISE = set("not use general field section cross medium auto hard customized".split())

_TOK = re.compile(r"[a-z]{3,}")


def grams(text: str) -> set[str]:
    """Lowercase -> word tokens (>=3 alpha, minus stopwords) -> unigrams +
    adjacent bigrams. The RUNTIME gate MUST use this exact tokenisation."""
    words = [w for w in _TOK.findall((text or "").lower()) if w not in STOPWORDS]
    out: set[str] = set(words)
    for i in range(len(words) - 1):
        out.add(words[i] + " " + words[i + 1])
    return out


def capability_tokens() -> set[str]:
    """Whole-word vocabulary of the capability reference (for guarding)."""
    if not _CAP_REF.exists():
        print(f"  [warn] {_CAP_REF} missing — guarded set will be empty")
        return set()
    return set(_TOK.findall(_CAP_REF.read_text().lower()))


def load_corpus() -> list[tuple[str, str, str, int]]:
    """Return (portal, bid_id, text, score) for every scored bid across tools."""
    rows: list[tuple[str, str, str, int]] = []
    for portal, (rel, tbl, textcol, pk) in SOURCES.items():
        path = _REPO_ROOT / rel
        if not path.exists():
            print(f"  [warn] {portal}: {path} missing — skipped")
            continue
        conn = sqlite3.connect(f"file:{path}?mode=ro", uri=True)
        sel = ", ".join(pk + [textcol, "pass1_score"])
        for r in conn.execute(f"SELECT {sel} FROM {tbl} WHERE pass1_score IS NOT NULL"):
            *pk_vals, text, score = r
            bid_id = "|".join("" if v is None else str(v) for v in pk_vals)
            rows.append((portal, bid_id, text or "", int(score)))
        conn.close()
    return rows


def export_corpus_excel(rows, out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "scored_corpus"
    ws.append(["portal", "bid_id", "score", "text"])
    for portal, bid_id, text, score in rows:
        ws.append([portal, bid_id, score, text])
    s2 = wb.create_sheet("summary")
    s2.append(["portal", "score", "count"])
    for (p, sc), n in sorted(Counter((p, sc) for p, _, _, sc in rows).items()):
        s2.append([p, sc, n])
    wb.save(out_path)


def _eliminates(g: set[str], phrases: set[str], words: set[str], guarded: set[str]):
    """Apply the two-tier gate to a bid's gram-set. Returns the matched keywords
    (sorted) if eliminated, else None. Mirrors the runtime gate exactly."""
    matched = (g & phrases) | (g & words)
    direct = (g & phrases) | (g & (words - guarded))
    if direct:
        return sorted(matched)
    # guarded-only path: need a second signal (>= 2 distinct unigram hits)
    if (g & guarded) and len(g & words) >= 2:
        return sorted(matched)
    return None


def mine(rows, cap_tokens):
    reject = [t for _, _, t, s in rows if s in (0, 1)]
    protect = [t for _, _, t, s in rows if s >= PROTECT_FROM]
    s2_texts = [t for _, _, t, s in rows if s == 2]

    def docfreq(corpus):
        c = Counter()
        for t in corpus:
            for x in grams(t):
                c[x] += 1
        return c

    rf, pf, s2f = docfreq(reject), docfreq(protect), docfreq(s2_texts)

    qualifying = {g for g, n in rf.items() if n >= MIN_SUPPORT and pf.get(g, 0) == 0}
    phrases = {g for g in qualifying if " " in g}
    unigrams = {g for g in qualifying if " " not in g}

    guarded_all = unigrams & cap_tokens
    guarded = guarded_all - GUARDED_NOISE           # active two-signal guards
    excluded = sorted(guarded_all & GUARDED_NOISE, key=lambda g: -rf[g])
    words = unigrams - set(excluded)                # active direct-eligible unigrams

    return phrases, words, guarded, excluded, rf, s2f


def export_score2_review(rows, phrases, words, guarded, out_path: Path) -> int:
    """Write the score-2 bids that WOULD be eliminated under the two-signal rule."""
    hits = []
    for portal, bid_id, text, score in rows:
        if score != 2:
            continue
        matched = _eliminates(grams(text), phrases, words, guarded)
        if matched:
            hits.append((portal, bid_id, text, matched))
    hits.sort(key=lambda h: -len(h[3]))
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "score2_review"
    ws.append(["portal", "bid_id", "text", "matched_keywords", "matched_count", "keep_decision"])
    for portal, bid_id, text, matched in hits:
        ws.append([portal, bid_id, text, ", ".join(matched), len(matched), ""])
    wb.save(out_path)
    return len(hits)


def main() -> int:
    runtime = Path(os.environ.get("BIDPLUS_RUNTIME_DIR", str(Path.home() / "bidplus-runtime")))
    corpus_xlsx = runtime / "exports" / "scored_corpus.xlsx"
    review_xlsx = runtime / "exports" / "score2_review.xlsx"

    print("Loading scored corpus from the three tool DBs…")
    rows = load_corpus()
    print(f"  {len(rows)} scored bids loaded.")
    export_corpus_excel(rows, corpus_xlsx)
    print(f"Wrote unified corpus: {corpus_xlsx}")

    cap = capability_tokens()
    phrases, words, guarded, excluded, rf, s2f = mine(rows, cap)

    # validation: coverage + collision re-check from the gate itself
    band = {k: [t for _, _, t, s in rows if s == k] for k in range(6)}
    reject_n = len(band[0]) + len(band[1])
    cut = {k: sum(_eliminates(grams(t), phrases, words, guarded) is not None for t in band[k]) for k in band}
    reject_cut = cut[0] + cut[1]
    protect_cut = cut[3] + cut[4] + cut[5]

    s2_count = export_score2_review(rows, phrases, words, guarded, review_xlsx)
    print(f"Wrote score-2 review: {review_xlsx}")

    active = phrases | words
    keep_s2 = {k: s2f[k] for k in sorted(active) if s2f.get(k, 0) > 0}

    payload = {
        "_meta": {
            "generated_at": _dt.datetime.now(_dt.UTC).isoformat(),
            "generator": "bidplus/scripts/mine_eliminators.py (protect>=3)",
            "thresholds": {"min_reject_support": MIN_SUPPORT,
                           "keep_guard": f"protect>=*{PROTECT_FROM}* (tolerate score-2, zero score-3/4/5)"},
            "sources": {p: rel for p, (rel, *_rest) in SOURCES.items()},
            "validation": {
                "phrases": len(phrases),
                "words": len(words),
                "guarded_two_signal": len(guarded),
                "excluded_generic_unigrams": len(excluded),
                "reject_0_1": reject_n,
                "reject_eliminated": reject_cut,
                "reject_eliminated_pct": round(100 * reject_cut / reject_n, 1) if reject_n else 0,
                "corpus_eliminated_pct": round(100 * (reject_cut + cut[2]) / len(rows), 1) if rows else 0,
                "score2_eliminated": cut[2],
                "score_3_4_5_collisions": protect_cut,
            },
            "matching": (
                "lowercase; tokenize [a-z]{3,} minus stopwords; unigrams + adjacent "
                "bigrams. Eliminate (score 0) on any phrase hit or any non-guarded word "
                "hit. A guarded word eliminates ONLY under the two-signal rule (a phrase "
                "hit or >=2 distinct unigram hits). The runtime gate MUST replicate grams()."
            ),
            "stopwords": sorted(STOPWORDS),
            "guarded_unigrams_two_signal": sorted(guarded),
            "excluded_generic_unigrams": excluded,
            "reject_support_per_keyword": {k: rf[k] for k in sorted(active)},
            "keep_s2_per_keyword": keep_s2,
            "review_status": (
                "MACHINE-GENERATED — review before enabling. Score-2 collisions are "
                "tolerated by design; eyeball score2_review.xlsx. Eliminated bids are "
                "SOFT-FLAGGED (score 0 + pass1_method='keyword' + pass1_eliminated_by), "
                "never hard-hidden; a human override quarantines the keyword from the next mine."
            ),
        },
        "phrases": sorted(phrases, key=lambda g: -rf[g]),
        "words": sorted(words, key=lambda g: -rf[g]),
    }
    _JSON_OUT.write_text(json.dumps(payload, indent=2, ensure_ascii=False))
    print(f"Wrote keyword list: {_JSON_OUT}")

    v = payload["_meta"]["validation"]
    print(
        f"\n  phrases={v['phrases']}  words={v['words']}  guarded={v['guarded_two_signal']}  "
        f"(trimmed noise held for review: {v['excluded_generic_unigrams']})\n"
        f"  reject eliminated {v['reject_eliminated']}/{v['reject_0_1']} "
        f"({v['reject_eliminated_pct']}% of reject, {v['corpus_eliminated_pct']}% of corpus)\n"
        f"  score-2 knowingly dropped: {v['score2_eliminated']}   "
        f"FALSE-ELIM score>=3: {v['score_3_4_5_collisions']}"
    )
    return 0 if v["score_3_4_5_collisions"] == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
